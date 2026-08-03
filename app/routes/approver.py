from datetime import datetime, timezone
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user

from app import db
from app.models import Ticket, ApprovalLog, ApprovalAction, TicketStatus, User, Notification, Scope, Category, IssueForm
from app.utils.email import send_ticket_approved, send_ticket_rejected, send_ticket_sent_back
from app.routes.auth import role_required

appr_bp = Blueprint('appr', __name__, url_prefix='/approver', template_folder='../templates/approver')


# Excel helpers
def sanitize_cell(value):
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@'):
        return "'" + value
    return value

def _export_build_workbook(tickets):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Requests'
    header_fill = PatternFill(start_color='01696f', end_color='01696f', fill_type='solid')
    header_font = Font(color='ffffff', bold=True, size=11)
    alt_fill = PatternFill(start_color='f7f6f2', end_color='f7f6f2', fill_type='solid')
    headers = ['Ticket #', 'Subject', 'Form Name', 'Scope', 'Raised By', 'Assigned To', 'Status', 'Created Date', 'Resolved Date']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for i, t in enumerate(tickets, start=2):
        row = [
            sanitize_cell(t.ticket_number),
            sanitize_cell(t.subject or ''),
            sanitize_cell(t.issue_form.name if t.issue_form else ''),
            sanitize_cell(t.scope or ''),
            sanitize_cell(t.creator.display_name if t.creator else ''),
            sanitize_cell(t.assignee.display_name if t.assignee else ''),
            sanitize_cell(t.current_status or ''),
            sanitize_cell(t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else ''),
            sanitize_cell(t.updated_at.strftime('%Y-%m-%d %H:%M') if t.updated_at and t.current_status in (
                TicketStatus.APPROVED.value,
                TicketStatus.REJECTED.value,
                TicketStatus.SENT_TO_FULFILMENT.value
            ) else '')
        ]
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=val)
            if (i - 2) % 2 == 1:
                from openpyxl.styles import PatternFill as _PF
                cell.fill = alt_fill
    # Auto-width
    for col in ws.columns:
        max_len = 15
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2
    return wb


def create_notification(user_id, title, message, link=None):
    notif = Notification(user_id=user_id, title=title, message=message, link=link)
    db.session.add(notif)


@appr_bp.route('/queue')
@login_required
@role_required('approver', 'admin')
def queue():
    """Default redirect to first active scope queue (or generic)."""
    first_scope = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).first()
    if first_scope:
        return redirect(url_for('appr.queue_dynamic', scope_name=first_scope.name))
    # Fallback: show all without scope filter
    tickets = _queue_by_scope(None)
    active_scopes = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).all()
    return render_template('approver/queue.html', tickets=tickets, queue_scope='All', active_scopes=active_scopes)


def _queue_by_scope(scope):
    """Helper to get active tickets for current approver, optionally filtered by scope."""
    from sqlalchemy import or_

    query = Ticket.query.filter(
        Ticket.assigned_to == current_user.id,
        Ticket.current_status.in_([
            TicketStatus.PENDING.value,
            TicketStatus.UNDER_REVIEW.value,
            TicketStatus.NEEDS_CLARIFICATION.value
        ])
    )

    if scope:
        # Show tickets with matching scope OR empty scope (unscoped forms)
        query = query.filter(or_(Ticket.scope == scope, Ticket.scope == ''))
    else:
        # For combined view, show all
        pass

    return query.order_by(Ticket.created_at.desc()).limit(500).all()


@appr_bp.route('/queue/ar')
@login_required
@role_required('approver', 'admin')
def queue_ar():
    return redirect(url_for('appr.queue_dynamic', scope_name='AR'))


@appr_bp.route('/queue/gl')
@login_required
@role_required('approver', 'admin')
def queue_gl():
    return redirect(url_for('appr.queue_dynamic', scope_name='GL'))


@appr_bp.route('/queue/<scope_name>')
@login_required
@role_required('approver', 'admin')
def queue_dynamic(scope_name):
    scope = (scope_name or '').upper()
    tickets = _queue_by_scope(scope)
    active_scopes = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).all()
    return render_template('approver/queue.html', tickets=tickets, queue_scope=scope, active_scopes=active_scopes)


@appr_bp.route('/history')
@login_required
@role_required('approver', 'admin')
def approval_history():
    resolved_tickets = Ticket.query.filter(
        Ticket.assigned_to == current_user.id,
        Ticket.current_status.in_([
            TicketStatus.APPROVED.value,
            TicketStatus.REJECTED.value,
            TicketStatus.SENT_TO_FULFILMENT.value
        ])
    ).order_by(Ticket.updated_at.desc()).limit(500).all()
    active_scopes = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).all()
    return render_template('approver/approval_history.html',
                           tickets=resolved_tickets, active_scopes=active_scopes)


@appr_bp.route('/request-history')
@login_required
@role_required('approver', 'admin')
def request_history():
    """All tickets across system with filters and pagination."""
    status = request.args.get('status', '').strip()
    scope = request.args.get('scope', '').strip()
    created_from = request.args.get('from', '').strip()
    created_to = request.args.get('to', '').strip()
    page = request.args.get('page', 1, type=int)

    q = Ticket.query
    if status:
        q = q.filter(Ticket.current_status == status)
    if scope:
        q = q.filter(Ticket.scope == scope)
    try:
        if created_from:
            fd = datetime.strptime(created_from, '%Y-%m-%d')
            q = q.filter(Ticket.created_at >= fd)
        if created_to:
            td = datetime.strptime(created_to, '%Y-%m-%d')
            from datetime import timedelta as _td
            q = q.filter(Ticket.created_at < td + _td(days=1))
    except Exception:
        pass
    pagination = q.order_by(Ticket.created_at.desc()).paginate(page=page, per_page=50, error_out=False)
    tickets = pagination.items
    active_scopes = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).all()
    return render_template('approver/request_history.html', tickets=tickets, pagination=pagination, active_scopes=active_scopes,
                           filters={'status': status, 'scope': scope, 'from': created_from, 'to': created_to})


@appr_bp.route('/analytics')
@login_required
@role_required('approver', 'admin')
def analytics():
    """Category-wise ticket breakdown, within scopes the approver has access to."""
    from sqlalchemy import func

    if current_user.role == 'approver':
        my_scopes = current_user.scopes.all()
    else:
        my_scopes = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).all()
    scope_ids = [s.id for s in my_scopes]

    selected_scope = request.args.get('scope', '').strip()
    selected_user = request.args.get('user_id', type=int)

    base_filters = [IssueForm.scope_id.in_(scope_ids)]
    if selected_scope:
        sc = Scope.query.filter_by(name=selected_scope).first()
        if sc:
            base_filters.append(IssueForm.scope_id == sc.id)
    if selected_user:
        base_filters.append(Ticket.created_by == selected_user)

    rows = db.session.query(Category.name, func.count(Ticket.id)) \
        .select_from(Ticket) \
        .join(IssueForm, Ticket.form_id == IssueForm.id) \
        .outerjoin(Category, IssueForm.category_id == Category.id) \
        .filter(*base_filters) \
        .group_by(Category.name) \
        .all()
    category_data = [{'label': (name or 'Uncategorized'), 'count': count} for name, count in rows]
    total_count = sum(c['count'] for c in category_data)

    users_list = db.session.query(User.id, User.display_name) \
        .join(Ticket, Ticket.created_by == User.id) \
        .join(IssueForm, Ticket.form_id == IssueForm.id) \
        .filter(IssueForm.scope_id.in_(scope_ids)) \
        .distinct().order_by(User.display_name.asc()).all()

    return render_template('approver/analytics.html',
                           scopes=my_scopes, category_data=category_data, total_count=total_count,
                           users_list=users_list, selected_scope=selected_scope, selected_user=selected_user)


@appr_bp.route('/history/export')
@login_required
@role_required('approver', 'admin')
def export_request_history():
    status = request.args.get('status', '').strip()
    scope = request.args.get('scope', '').strip()
    created_from = request.args.get('from', '').strip()
    created_to = request.args.get('to', '').strip()

    q = Ticket.query
    if status:
        q = q.filter(Ticket.current_status == status)
    if scope:
        q = q.filter(Ticket.scope == scope)
    try:
        if created_from:
            fd = datetime.strptime(created_from, '%Y-%m-%d')
            q = q.filter(Ticket.created_at >= fd)
        if created_to:
            td = datetime.strptime(created_to, '%Y-%m-%d')
            from datetime import timedelta as _td
            q = q.filter(Ticket.created_at < td + _td(days=1))
    except Exception:
        pass
    tickets = q.order_by(Ticket.created_at.desc()).all()
    wb = _export_build_workbook(tickets)
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='approver_request_history.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@appr_bp.route('/approval-history/export')
@login_required
@role_required('approver', 'admin')
def export_approval_history():
    status = request.args.get('status', '').strip()
    scope = request.args.get('scope', '').strip()
    created_from = request.args.get('from', '').strip()
    created_to = request.args.get('to', '').strip()

    q = Ticket.query.filter(
        Ticket.assigned_to == current_user.id,
        Ticket.current_status.in_([
            TicketStatus.APPROVED.value,
            TicketStatus.REJECTED.value,
            TicketStatus.SENT_TO_FULFILMENT.value
        ])
    )
    if status:
        q = q.filter(Ticket.current_status == status)
    if scope:
        q = q.filter(Ticket.scope == scope)
    try:
        if created_from:
            fd = datetime.strptime(created_from, '%Y-%m-%d')
            q = q.filter(Ticket.updated_at >= fd)
        if created_to:
            td = datetime.strptime(created_to, '%Y-%m-%d')
            from datetime import timedelta as _td
            q = q.filter(Ticket.updated_at < td + _td(days=1))
    except Exception:
        pass
    tickets = q.order_by(Ticket.updated_at.desc()).all()
    wb = _export_build_workbook(tickets)
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='approver_approval_history.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@appr_bp.route('/ticket/<int:ticket_id>')
@login_required
@role_required('approver', 'admin')
def ticket_detail(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)

    if ticket.assigned_to != current_user.id and current_user.role != 'admin':
        flash('This ticket is not assigned to you.', 'error')
        return redirect(url_for('appr.queue'))

    logs = ApprovalLog.query.filter_by(ticket_id=ticket.id).order_by(ApprovalLog.timestamp.desc()).all()
    approvers = User.query.filter(
        User.role == 'approver', User.is_active == True,
        User.id != current_user.id
    ).all()
    return render_template('approver/ticket_detail.html', ticket=ticket, logs=logs, approvers=approvers)


@appr_bp.route('/ticket/<int:ticket_id>/approve', methods=['POST'])
@login_required
@role_required('approver', 'admin')
def approve_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.assigned_to != current_user.id and current_user.role != 'admin':
        flash('This ticket is not assigned to you.', 'error')
        return redirect(url_for('appr.queue'))

    if ticket.current_status not in ['Pending', 'Under Review']:
        flash('This ticket cannot be approved in its current state.', 'error')
        return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))

    comment = request.form.get('comment', '').strip()
    if not comment:
        flash('Please provide a comment for this approval.', 'error')
        return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))

    now = datetime.now(timezone.utc)

    ticket.current_status = TicketStatus.SENT_TO_FULFILMENT.value
    ticket.updated_at = now

    log_approved = ApprovalLog(
        ticket_id=ticket.id, action_by=current_user.id,
        action=ApprovalAction.APPROVED.value, comment=comment,
        timestamp=now
    )
    log_fulfilment = ApprovalLog(
        ticket_id=ticket.id, action_by=current_user.id,
        action=ApprovalAction.SENT_TO_FULFILMENT.value,
        comment='Automatically transitioned after approval',
        timestamp=now
    )
    db.session.add(log_approved)
    db.session.add(log_fulfilment)

    create_notification(ticket.created_by, 'Ticket Approved',
                       f'Your ticket {ticket.ticket_number} has been approved',
                       url_for('req.ticket_detail', ticket_id=ticket.id))

    db.session.commit()
    req_url = url_for('req.ticket_detail', ticket_id=ticket.id, _external=True)
    try:
        send_ticket_approved(ticket, req_url)
    except Exception:
        import logging
        logging.exception('Failed to send approval notification email')
    flash(f'Ticket {ticket.ticket_number} approved and sent to fulfilment.', 'success')
    return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))


@appr_bp.route('/ticket/<int:ticket_id>/reject', methods=['POST'])
@login_required
@role_required('approver', 'admin')
def reject_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.assigned_to != current_user.id and current_user.role != 'admin':
        flash('This ticket is not assigned to you.', 'error')
        return redirect(url_for('appr.queue'))

    if ticket.current_status not in ['Pending', 'Under Review']:
        flash('This ticket cannot be rejected in its current state.', 'error')
        return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))

    comment = request.form.get('comment', '').strip()
    if not comment:
        flash('Please provide a reason for rejection.', 'error')
        return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))

    ticket.current_status = TicketStatus.REJECTED.value
    ticket.updated_at = datetime.now(timezone.utc)
    log = ApprovalLog(ticket_id=ticket.id, action_by=current_user.id, action=ApprovalAction.REJECTED.value, comment=comment)
    db.session.add(log)

    create_notification(ticket.created_by, 'Ticket Rejected', f'Your ticket {ticket.ticket_number} has been rejected', url_for('req.ticket_detail', ticket_id=ticket.id))

    db.session.commit()
    req_url = url_for('req.ticket_detail', ticket_id=ticket.id, _external=True)
    try:
        send_ticket_rejected(ticket, comment, req_url)
    except Exception:
        import logging
        logging.exception('Failed to send rejection notification email')
    flash(f'Ticket {ticket.ticket_number} rejected.', 'error')
    return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))


@appr_bp.route('/ticket/<int:ticket_id>/send-back', methods=['POST'])
@login_required
@role_required('approver', 'admin')
def send_back_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.assigned_to != current_user.id and current_user.role != 'admin':
        flash('This ticket is not assigned to you.', 'error')
        return redirect(url_for('appr.queue'))

    if ticket.current_status not in ['Pending', 'Under Review']:
        flash('This ticket cannot be sent back in its current state.', 'error')
        return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))

    comment = request.form.get('comment', '').strip()
    if not comment:
        flash('Please specify what clarification is needed.', 'error')
        return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))

    ticket.current_status = TicketStatus.NEEDS_CLARIFICATION.value
    ticket.updated_at = datetime.now(timezone.utc)
    log = ApprovalLog(ticket_id=ticket.id, action_by=current_user.id, action=ApprovalAction.SENT_BACK.value, comment=comment)
    db.session.add(log)

    create_notification(ticket.created_by, 'Clarification Needed', f'Clarification needed on {ticket.ticket_number}', url_for('req.ticket_detail', ticket_id=ticket.id))

    db.session.commit()
    req_url = url_for('req.ticket_detail', ticket_id=ticket.id, _external=True)
    try:
        send_ticket_sent_back(ticket, comment, req_url)
    except Exception:
        import logging
        logging.exception('Failed to send sent-back notification email')
    flash(f'Ticket {ticket.ticket_number} sent back for clarification.', 'info')
    return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))


@appr_bp.route('/ticket/<int:ticket_id>/reassign', methods=['POST'])
@login_required
@role_required('approver', 'admin')
def reassign_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.assigned_to != current_user.id and current_user.role != 'admin':
        flash('This ticket is not assigned to you.', 'error')
        return redirect(url_for('appr.queue'))

    new_assignee_id = request.form.get('new_assignee_id', type=int)
    reason = request.form.get('comment', '').strip()

    if not new_assignee_id or not reason:
        flash('Please select a new approver and provide a reason.', 'error')
        return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))

    new_assignee = User.query.get(new_assignee_id)
    if not new_assignee or new_assignee.role != 'approver' or not new_assignee.is_active:
        flash('Invalid approver selected.', 'error')
        return redirect(url_for('appr.ticket_detail', ticket_id=ticket.id))

    old_assignee_name = ticket.assignee.display_name if ticket.assignee else 'Unassigned'
    ticket.assigned_to = new_assignee_id
    ticket.updated_at = datetime.now(timezone.utc)

    log = ApprovalLog(
        ticket_id=ticket.id, action_by=current_user.id,
        action=ApprovalAction.REASSIGNED.value,
        comment=f'{reason} (Reassigned from {old_assignee_name} to {new_assignee.display_name})'
    )
    db.session.add(log)

    create_notification(new_assignee_id, 'Ticket Reassigned', f'Ticket {ticket.ticket_number} has been assigned to you', url_for('appr.ticket_detail', ticket_id=ticket.id))
    create_notification(ticket.created_by, 'Ticket Reassigned', f'Your ticket {ticket.ticket_number} has been reassigned', url_for('req.ticket_detail', ticket_id=ticket.id))

    db.session.commit()

    flash(f'Ticket reassigned to {new_assignee.display_name}.', 'success')
    return redirect(url_for('appr.queue'))


@appr_bp.route('/bulk-action', methods=['POST'])
@login_required
@role_required('approver', 'admin')
def bulk_action():
    data = request.get_json(silent=True) or {}
    ticket_ids = data.get('ticket_ids', [])
    action = data.get('action', '')
    comment = data.get('comment', '')

    if not ticket_ids or action not in ['approve', 'reject']:
        return jsonify({'success': False, 'error': 'Invalid request'}), 400

    processed = 0
    for tid in ticket_ids:
        ticket = Ticket.query.get(tid)
        if not ticket or ticket.assigned_to != current_user.id:
            continue
        if ticket.current_status not in ['Pending', 'Under Review']:
            continue

        now = datetime.now(timezone.utc)

        if action == 'approve':
            ticket.current_status = TicketStatus.SENT_TO_FULFILMENT.value
            ticket.updated_at = now
            action_name = ApprovalAction.APPROVED.value
            log_comment = comment or f'Bulk approved by {current_user.display_name}'
            create_notification(ticket.created_by, 'Ticket Approved', f'Your ticket {ticket.ticket_number} has been approved', url_for('req.ticket_detail', ticket_id=ticket.id))
            req_url = url_for('req.ticket_detail', ticket_id=ticket.id, _external=True)
            try:
                send_ticket_approved(ticket, req_url)
            except Exception:
                import logging
                logging.exception('Failed to send bulk approval email')
        else:
            ticket.current_status = TicketStatus.REJECTED.value
            ticket.updated_at = now
            action_name = ApprovalAction.REJECTED.value
            log_comment = comment or f'Bulk rejected by {current_user.display_name}'
            create_notification(ticket.created_by, 'Ticket Rejected', f'Your ticket {ticket.ticket_number} has been rejected', url_for('req.ticket_detail', ticket_id=ticket.id))
            req_url = url_for('req.ticket_detail', ticket_id=ticket.id, _external=True)
            try:
                send_ticket_rejected(ticket, comment, req_url)
            except Exception:
                import logging
                logging.exception('Failed to send bulk rejection email')

        log = ApprovalLog(ticket_id=ticket.id, action_by=current_user.id, action=action_name, comment=log_comment)
        db.session.add(log)
        processed += 1

    db.session.commit()
    return jsonify({'success': True, 'processed': processed})