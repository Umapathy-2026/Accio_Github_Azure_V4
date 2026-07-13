from datetime import datetime, timezone, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify, current_app
from flask_login import login_required, current_user
from werkzeug.security import generate_password_hash
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from app import db
from app.models import (
    User,
    IssueForm,
    Ticket,
    ApprovalLog,
    Notification,
    TicketStatus,
    ApprovalAction,
    AdminAuditLog,
    Scope,
    Category,
    UserScope,
    DEFAULT_FORM_FIELDS,
)
import re
from app.routes.auth import role_required

admin_bp = Blueprint('admin', __name__, url_prefix='/admin', template_folder='../templates/admin')


def log_admin_action(action, target_type=None, target_id=None, details=None):
    entry = AdminAuditLog(
        performed_by=current_user.id,
        action=action,
        target_type=target_type,
        target_id=target_id,
        details=details or {},
        ip_address=request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0].strip(),
        timestamp=datetime.now(timezone.utc)
    )
    db.session.add(entry)


@admin_bp.route('/dashboard')
@login_required
@role_required('admin')
def dashboard():
    now = datetime.now(timezone.utc)
    first_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    total_tickets = Ticket.query.count()
    total_this_month = Ticket.query.filter(Ticket.created_at >= first_of_month).count()
    pending = Ticket.query.filter_by(current_status='Pending').count()
    approved_this_month = Ticket.query.filter(
        Ticket.current_status == 'Sent to Fulfilment',
        Ticket.created_at >= first_of_month
    ).count()
    rejected_this_month = Ticket.query.filter(
        Ticket.current_status == 'Rejected',
        Ticket.created_at >= first_of_month
    ).count()

    recent_tickets = Ticket.query.order_by(Ticket.created_at.desc()).limit(10).all()

    return render_template('admin/dashboard.html',
                         total_tickets=total_tickets,
                         total_this_month=total_this_month,
                         pending=pending,
                         approved_this_month=approved_this_month,
                         rejected_this_month=rejected_this_month,
                         recent_tickets=recent_tickets)


@admin_bp.route('/tickets')
@login_required
@role_required('admin')
def all_tickets():
    tickets = Ticket.query.order_by(Ticket.created_at.desc()).limit(500).all()
    raw_count = Ticket.query.count()
    cap_hit = raw_count > 500
    users = User.query.filter_by(is_active=True).all()
    forms = IssueForm.query.filter_by(is_active=True).all()
    active_scopes = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).all()
    return render_template('admin/all_tickets.html',
                           tickets=tickets, users=users, forms=forms,
                           cap_hit=cap_hit, total=raw_count, active_scopes=active_scopes)


@admin_bp.route('/users', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def users():
    if request.method == 'POST':
        display_name = request.form.get('display_name', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        role = request.form.get('role', 'user')
        approver_id = request.form.get('approver_id', type=int)
        # Multi-scope assignment: list of ints
        scope_ids = request.form.getlist('scope_ids')
        try:
            scope_ids = [int(sid) for sid in scope_ids if sid]
        except Exception:
            scope_ids = []

        if not display_name or not email or not password:
            flash('Name, email, and password are required.', 'error')
            return redirect(url_for('admin.users'))

        if User.query.filter_by(email=email).first():
            flash('A user with this email already exists.', 'error')
            return redirect(url_for('admin.users'))

        from app.routes.auth import validate_password_strength
        errors = validate_password_strength(password)
        if errors:
            for err in errors:
                flash(err, 'error')
            return redirect(url_for('admin.users'))

        if approver_id:
            approver = User.query.filter_by(id=approver_id, is_active=True).first()
            if not approver or approver.role not in ('approver', 'admin'):
                flash('Invalid approver assignment. Approver must be an active approver or admin.', 'error')
                return redirect(url_for('admin.users'))

        user = User(
            email=email,
            display_name=display_name,
            password_hash=generate_password_hash(password),
            role=role,
            approver_id=approver_id if approver_id else None,
            is_active=True
        )
        db.session.add(user)
        db.session.flush()

        # Create UserScope rows for checked scopes (active scopes only)
        if scope_ids:
            active_scopes = {s.id for s in Scope.query.filter(Scope.id.in_(scope_ids), Scope.is_active == True).all()}
            for sid in active_scopes:
                db.session.add(UserScope(user_id=user.id, scope_id=sid))

        log_admin_action('USER_CREATED', 'user', user.id,
                         {'email': email, 'role': role, 'approver_id': approver_id, 'scope_ids': scope_ids})
        db.session.commit()
        flash(f'User {display_name} created successfully.', 'success')
        return redirect(url_for('admin.users'))

    users = User.query.order_by(User.created_at.desc()).all()
    approvers = User.query.filter_by(role='approver', is_active=True).all()
    active_scopes = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).all()
    return render_template('admin/users.html', users=users, approvers=approvers, scopes=active_scopes, now=datetime.now(timezone.utc))


@admin_bp.route('/users/<int:user_id>/edit', methods=['POST'])
@login_required
@role_required('admin')
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    old_role = user.role

    display_name = request.form.get('display_name', '').strip()
    role = request.form.get('role', 'user')
    approver_id = request.form.get('approver_id', type=int)
    is_active = request.form.get('is_active') == 'on'
    # Multi-scope selection
    scope_ids = request.form.getlist('scope_ids')
    try:
        scope_ids = [int(sid) for sid in scope_ids if sid]
    except Exception:
        scope_ids = []

    if display_name:
        user.display_name = display_name
    user.role = role
    user.approver_id = approver_id if approver_id else None
    user.is_active = is_active

    password = request.form.get('password', '')
    if password:
        from app.routes.auth import validate_password_strength
        pw_errors = validate_password_strength(password)
        if pw_errors:
            for err in pw_errors:
                flash(err, 'error')
            return redirect(url_for('admin.users'))
        user.password_hash = generate_password_hash(password)

    details = {'role_changed': old_role != role}
    if old_role != role:
        details['from_role'] = old_role
        details['to_role'] = role
    # Update user scopes: remove existing and insert new
    from sqlalchemy import delete
    db.session.execute(delete(UserScope).where(UserScope.user_id == user.id))
    if scope_ids:
        active_scopes = {s.id for s in Scope.query.filter(Scope.id.in_(scope_ids), Scope.is_active == True).all()}
        for sid in active_scopes:
            db.session.add(UserScope(user_id=user.id, scope_id=sid))
    details['scope_ids'] = scope_ids
    log_admin_action('USER_UPDATED', 'user', user.id, details)
    db.session.commit()
    flash(f'User {user.display_name} updated successfully.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:user_id>/unlock', methods=['POST'])
@login_required
@role_required('admin')
def unlock_user(user_id):
    user = User.query.get_or_404(user_id)
    user.failed_attempts = 0
    user.locked_until = None
    log_admin_action('USER_UNLOCKED', 'user', user.id)
    db.session.commit()
    flash(f'User {user.display_name} has been unlocked.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:user_id>/deactivate-check', methods=['GET'])
@login_required
@role_required('admin')
def deactivate_user_check(user_id):
    user = User.query.get_or_404(user_id)

    open_status_list = [TicketStatus.PENDING.value, TicketStatus.UNDER_REVIEW.value, TicketStatus.NEEDS_CLARIFICATION.value]

    created_open = Ticket.query.filter(
        Ticket.created_by == user_id,
        Ticket.current_status.in_(open_status_list)
    ).all()

    assigned_open = Ticket.query.filter(
        Ticket.assigned_to == user_id,
        Ticket.current_status.in_(open_status_list)
    ).all()

    total_open = len(created_open) + len(assigned_open)
    approvers = User.query.filter_by(role='approver', is_active=True).filter(User.id != user_id).all()
    # Subordinates of this user (active)
    subordinates = User.query.filter_by(approver_id=user_id, is_active=True).all()

    return jsonify({
        'has_dependencies': (total_open > 0) or (len(subordinates) > 0),
        'open_ticket_count': total_open,
        'created_count': len(created_open),
        'assigned_count': len(assigned_open),
        'created_ticket_ids': [t.id for t in created_open],
        'assigned_ticket_ids': [t.id for t in assigned_open],
        'user_id': user_id,
        'role': user.role,
        'subordinates_count': len(subordinates),
        'subordinates': [{'id': u.id, 'display_name': u.display_name, 'email': u.email} for u in subordinates],
        'approvers': [{'id': a.id, 'name': a.display_name} for a in approvers]
    })


@admin_bp.route('/users/<int:user_id>/deactivate', methods=['POST'])
@login_required
@role_required('admin')
def deactivate_user(user_id):
    user = User.query.get_or_404(user_id)
    reassign_created_to = request.form.get('reassign_to', type=int)
    reassign_assigned_to = request.form.get('reassign_assigned_to', type=int)
    reassign_users_to = request.form.get('reassign_users_to', type=int)

    open_status_list = [TicketStatus.PENDING.value, TicketStatus.UNDER_REVIEW.value, TicketStatus.NEEDS_CLARIFICATION.value]

    created_open = Ticket.query.filter(
        Ticket.created_by == user_id,
        Ticket.current_status.in_(open_status_list)
    ).all()

    assigned_open = Ticket.query.filter(
        Ticket.assigned_to == user_id,
        Ticket.current_status.in_(open_status_list)
    ).all()

    reassigned_count = 0
    reassigned_users = 0

    # If deactivating an approver, we must handle subordinate reassignment first
    if user.role == 'approver':
        if reassign_users_to is None:
            flash('You must select a new approver to reassign subordinates.', 'error')
            return redirect(url_for('admin.users'))
        new_mgr = User.query.filter_by(id=reassign_users_to, role='approver', is_active=True).first()
        if not new_mgr:
            flash('Invalid approver selected for subordinate reassignment.', 'error')
            return redirect(url_for('admin.users'))
        # Reassign all active users who have this approver
        subs = User.query.filter_by(approver_id=user_id, is_active=True).all()
        for s in subs:
            s.approver_id = new_mgr.id
            reassigned_users += 1
        if reassigned_users > 0:
            log_admin_action('USERS_REASSIGNED_FROM_APPROVER', 'user', user.id,
                             {'from': user.display_name, 'to': new_mgr.display_name, 'count': reassigned_users})

    if created_open and reassign_created_to:
        new_assignee = User.query.filter_by(
            id=reassign_created_to, role='approver', is_active=True
        ).first()
        if new_assignee:
            for ticket in created_open:
                ticket.assigned_to = reassign_created_to
                log = ApprovalLog(
                    ticket_id=ticket.id, action_by=current_user.id,
                    action=ApprovalAction.REASSIGNED.value,
                    comment=f'Auto-reassigned from {user.display_name} (deactivated) to {new_assignee.display_name}'
                )
                db.session.add(log)
                reassigned_count += 1

    if assigned_open and reassign_assigned_to:
        new_approver = User.query.filter_by(
            id=reassign_assigned_to, role='approver', is_active=True
        ).first()
        if new_approver:
            for ticket in assigned_open:
                ticket.assigned_to = reassign_assigned_to
                log = ApprovalLog(
                    ticket_id=ticket.id, action_by=current_user.id,
                    action=ApprovalAction.REASSIGNED.value,
                    comment=f'Approver queue transferred from {user.display_name} (deactivated) to {new_approver.display_name}'
                )
                db.session.add(log)
                reassigned_count += 1

    user.is_active = False
    log_admin_action('USER_DEACTIVATED', 'user', user.id,
                    {'tickets_reassigned': reassigned_count,
                     'created_tickets': len(created_open),
                     'assigned_tickets': len(assigned_open),
                     'users_reassigned': reassigned_users})
    db.session.commit()

    if reassigned_count > 0:
        flash(f'User deactivated. {reassigned_count} tickets reassigned.', 'success')
    else:
        flash(f'User {user.display_name} deactivated.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/forms')
@login_required
@role_required('admin')
def forms():
    all_forms = IssueForm.query.filter_by(is_deleted=False).order_by(IssueForm.created_at.desc()).all()
    scopes = Scope.query.order_by(Scope.name.asc()).all()
    # Preload categories (active + inactive) to display names; creation dropdown will use active only
    categories = Category.query.order_by(Category.name.asc()).all()
    active_scopes = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).all()
    active_categories = Category.query.filter_by(is_active=True).order_by(Category.name.asc()).all()
    return render_template('admin/forms.html', forms=all_forms, scopes=scopes, categories=categories, active_scopes=active_scopes, active_categories=active_categories)


@admin_bp.route('/forms/<int:form_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def toggle_form(form_id):
    form = IssueForm.query.get_or_404(form_id)
    form.is_active = not form.is_active
    log_admin_action('FORM_TOGGLED', 'form', form.id, {'is_active': form.is_active, 'form_name': form.name})
    db.session.commit()
    flash(f'Form "{form.name}" {"activated" if form.is_active else "deactivated"}.', 'success')
    return redirect(url_for('admin.forms'))


@admin_bp.route('/forms/create', methods=['POST'])
@login_required
@role_required('admin')
def create_form():
    name = request.form.get('form_name', '').strip()
    scope_id = request.form.get('scope_id', type=int)
    category_id = request.form.get('category_id', type=int)

    if not name:
        flash('Form name is required.', 'error')
        return redirect(url_for('admin.forms'))

    # Forms are unique by name among non-deleted ones
    if IssueForm.query.filter(IssueForm.name == name, IssueForm.is_deleted == False).first():
        flash('A form with this name already exists.', 'error')
        return redirect(url_for('admin.forms'))

    # Validate scope
    scope_obj = Scope.query.filter_by(id=scope_id, is_active=True).first()
    if not scope_obj:
        flash('Invalid scope selected.', 'error')
        return redirect(url_for('admin.forms'))

    cat_obj = None
    if category_id:
        cat_obj = Category.query.filter_by(id=category_id, is_active=True).first()
        if not cat_obj or cat_obj.scope_id != scope_id:
            flash('Invalid category selected for the chosen scope.', 'error')
            return redirect(url_for('admin.forms'))

    from copy import deepcopy
    default_fields = deepcopy(DEFAULT_FORM_FIELDS)
    new_form = IssueForm(name=name, fields=default_fields, is_active=True, scope_id=scope_id, category_id=category_id)
    db.session.add(new_form)
    db.session.flush()
    log_admin_action('FORM_CREATED', 'form', new_form.id, {'form_name': name, 'scope_id': scope_id, 'category_id': category_id})
    db.session.commit()
    flash(f'Form "{name}" created successfully.', 'success')
    return redirect(url_for('admin.edit_form', form_id=new_form.id))


@admin_bp.route('/forms/<int:form_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def edit_form(form_id):
    form = IssueForm.query.get_or_404(form_id)

    if request.method == 'POST':
        form_name = request.form.get('form_name', '').strip()
        fields_json = request.form.get('fields_json', '[]')

        if not form_name:
            flash('Form name is required.', 'error')
            return redirect(url_for('admin.edit_form', form_id=form_id))
        
        # Check if name changed and new name doesn't exist
        if form_name != form.name and IssueForm.query.filter(IssueForm.name == form_name, IssueForm.is_deleted == False).first():
            flash('A form with this name already exists.', 'error')
            return redirect(url_for('admin.edit_form', form_id=form_id))
        
        import json
        try:
            fields = json.loads(fields_json)
            # Enforce protected fields and ordering
            def ensure_protected(current_fields):
                # Build custom fields only from non-protected entries
                protected_ids = {'subject', 'description', 'attachments'}
                custom = [f for f in current_fields if str(f.get('id') or f.get('name', '')).lower() not in protected_ids and not f.get('protected', False)]
                # Standardize custom field keys
                for f in custom:
                    # normalized structure: id/name/label/type/required/options
                    if 'id' not in f and 'name' in f:
                        f['id'] = f['name']
                    f.setdefault('required', False)
                    f.setdefault('type', 'text')
                # Return rebuilt fields list
                return [
                    {"id": "subject", "label": "Subject", "type": "text", "required": True,  "protected": True},
                    *custom,
                    {"id": "description", "label": "Description", "type": "textarea", "required": False, "protected": True},
                    {"id": "attachments", "label": "Attachments", "type": "file", "required": False, "protected": True},
                ]

            rebuilt = ensure_protected(fields)
            form.name = form_name
            form.fields = rebuilt
            log_admin_action('FORM_UPDATED', 'form', form.id,
                             {'form_name': form.name, 'field_count': len(rebuilt)})
            db.session.commit()
            flash(f'Form "{form.name}" updated successfully.', 'success')
        except json.JSONDecodeError:
            flash('Invalid fields data.', 'error')

        return redirect(url_for('admin.forms'))

    return render_template('admin/edit_form.html', form=form)

@admin_bp.route('/forms/<int:form_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def delete_form(form_id):
    form = IssueForm.query.get_or_404(form_id)
    form.is_deleted = True
    form.is_active = False
    log_admin_action('FORM_DELETED', 'form', form.id, {'form_name': form.name})
    db.session.commit()
    flash(f'Form "{form.name}" deleted.', 'success')
    return redirect(url_for('admin.forms'))

############################
# Scopes Management
############################

@admin_bp.route('/scopes', methods=['GET'])
@login_required
@role_required('admin')
def scopes():
    all_scopes = Scope.query.order_by(Scope.created_at.desc()).all()
    return render_template('admin/scopes.html', scopes=all_scopes)

@admin_bp.route('/scopes/create', methods=['POST'])
@login_required
@role_required('admin')
def create_scope():
    name = (request.form.get('name') or '').strip()
    if not name:
        flash('Scope name is required.', 'error')
        return redirect(url_for('admin.scopes'))
    # Enforce uniqueness case-insensitively
    from sqlalchemy import func
    existing = Scope.query.filter(func.lower(Scope.name) == name.lower()).first()
    if existing:
        flash('A scope with this name already exists.', 'error')
        return redirect(url_for('admin.scopes'))
    scope = Scope(name=name, is_active=True)
    db.session.add(scope)
    db.session.flush()
    log_admin_action('SCOPE_CREATED', 'scope', scope.id, {'name': name})
    db.session.commit()
    flash(f'Scope "{name}" created successfully.', 'success')
    return redirect(url_for('admin.scopes'))

@admin_bp.route('/scopes/<int:scope_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def toggle_scope(scope_id):
    scope = Scope.query.get_or_404(scope_id)
    # If deactivating, block if any active forms exist under this scope
    if scope.is_active:
        active_forms = IssueForm.query.filter_by(scope_id=scope.id, is_active=True, is_deleted=False).count()
        if active_forms > 0:
            flash(f'{active_forms} forms are active under this scope. Deactivate those forms first.', 'error')
            return redirect(url_for('admin.scopes'))
    scope.is_active = not scope.is_active
    log_admin_action('SCOPE_TOGGLED', 'scope', scope.id, {'is_active': scope.is_active})
    db.session.commit()
    flash(f'Scope "{scope.name}" {"activated" if scope.is_active else "deactivated"}.', 'success')
    return redirect(url_for('admin.scopes'))

############################
# Categories Management
############################

@admin_bp.route('/categories', methods=['GET'])
@login_required
@role_required('admin')
def categories():
    # Group by scope
    scopes = Scope.query.order_by(Scope.name.asc()).all()
    # Eager-load categories for template grouping
    cats = Category.query.order_by(Category.scope_id.asc(), Category.name.asc()).all()
    # Active scopes for creation dropdown
    active_scopes = Scope.query.filter_by(is_active=True).order_by(Scope.name.asc()).all()
    return render_template('admin/categories.html', scopes=scopes, categories=cats, active_scopes=active_scopes)

@admin_bp.route('/categories/create', methods=['POST'])
@login_required
@role_required('admin')
def create_category():
    name = (request.form.get('name') or '').strip()
    scope_id = request.form.get('scope_id', type=int)
    if not name or not scope_id:
        flash('Category name and scope are required.', 'error')
        return redirect(url_for('admin.categories'))
    scope = Scope.query.filter_by(id=scope_id, is_active=True).first()
    if not scope:
        flash('Invalid scope selected.', 'error')
        return redirect(url_for('admin.categories'))
    # Enforce unique per scope
    from sqlalchemy import func
    existing = Category.query.filter(
        Category.scope_id == scope_id,
        func.lower(Category.name) == name.lower()
    ).first()
    if existing:
        flash('A category with this name already exists in the selected scope.', 'error')
        return redirect(url_for('admin.categories'))
    cat = Category(name=name, scope_id=scope_id, is_active=True)
    db.session.add(cat)
    db.session.flush()
    log_admin_action('CATEGORY_CREATED', 'category', cat.id, {'name': name, 'scope_id': scope_id})
    db.session.commit()
    flash(f'Category "{name}" created successfully.', 'success')
    return redirect(url_for('admin.categories'))

@admin_bp.route('/categories/<int:category_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def toggle_category(category_id):
    cat = Category.query.get_or_404(category_id)
    # If deactivating, block if any active forms exist under this category
    if cat.is_active:
        active_forms = IssueForm.query.filter_by(category_id=cat.id, is_active=True, is_deleted=False).count()
        if active_forms > 0:
            flash(f'{active_forms} forms are active under this category. Deactivate those forms first.', 'error')
            return redirect(url_for('admin.categories'))
    cat.is_active = not cat.is_active
    log_admin_action('CATEGORY_TOGGLED', 'category', cat.id, {'is_active': cat.is_active})
    db.session.commit()
    flash(f'Category "{cat.name}" {"activated" if cat.is_active else "deactivated"}.', 'success')
    return redirect(url_for('admin.categories'))

def _sanitize_filename_part(value, default='all'):
    return re.sub(r'[^a-zA-Z0-9_\-]', '', str(value or default))[:30]

@admin_bp.route('/tickets/export')
@login_required
@role_required('admin')
def export_tickets_standard():
    """Export all tickets with standardized columns: Ticket#, Subject, Form Name,
    Scope, Raised By, Assigned To, Status, Created Date, Resolved Date.
    Applies same filters as all_tickets view: status, scope, date range.
    """
    from app.routes.requester import _export_build_workbook, sanitize_cell
    from app.models import TicketStatus

    status = request.args.get('status', '').strip()
    scope = request.args.get('scope', '').strip()
    created_from = request.args.get('from', '').strip()
    created_to = request.args.get('to', '').strip()

    q = Ticket.query
    if status:
        q = q.filter(Ticket.current_status == status)
    if scope:
        q = q.filter(Ticket.scope == scope)
    try:
        if created_from:
            fd = datetime.strptime(created_from, '%Y-%m-%d')
            q = q.filter(Ticket.created_at >= fd)
        if created_to:
            td = datetime.strptime(created_to, '%Y-%m-%d')
            q = q.filter(Ticket.created_at < td + timedelta(days=1))
    except Exception:
        pass

    tickets = q.order_by(Ticket.created_at.desc()).all()

    log_admin_action('EXPORT_TRIGGERED', 'ticket', None,
                     {'view': 'all_tickets', 'status': status, 'scope': scope,
                      'from': created_from, 'to': created_to, 'count': len(tickets)})

    wb = _export_build_workbook(tickets)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='admin_all_tickets.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/audit-log')
@login_required
@role_required('admin')
def audit_log():
    page = request.args.get('page', 1, type=int)
    logs = AdminAuditLog.query.order_by(
        AdminAuditLog.timestamp.desc()
    ).paginate(page=page, per_page=50, error_out=False)
    return render_template('admin/audit_log.html', logs=logs)


@admin_bp.route('/export-audit')
@login_required
@role_required('admin')
def export_audit():
    search = request.args.get('search', '').strip().lower()
    action_filter = request.args.get('action', '').strip()

    query = AdminAuditLog.query
    if search:
        from sqlalchemy import cast, String
        query = query.filter(
            db.or_(
                AdminAuditLog.action.ilike(f'%{search}%'),
                AdminAuditLog.target_type.ilike(f'%{search}%'),
                cast(AdminAuditLog.target_id, String).ilike(f'%{search}%')
            )
        )
    if action_filter:
        query = query.filter(AdminAuditLog.action == action_filter)

    logs = query.order_by(AdminAuditLog.timestamp.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Audit Log Export'

    header_fill = PatternFill(start_color='01696f', end_color='01696f', fill_type='solid')
    header_font = Font(color='ffffff', bold=True, size=11)
    alt_fill = PatternFill(start_color='f7f6f2', end_color='f7f6f2', fill_type='solid')

    headers = ['Timestamp', 'Admin', 'Action', 'Target Type', 'Target ID', 'IP Address', 'Details']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, log in enumerate(logs, 2):
        row_data = [
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S') if log.timestamp else '',
            log.actor.display_name if log.actor else '—',
            log.action,
            log.target_type or '—',
            log.target_id or '—',
            log.ip_address or '—',
            str(log.details) if log.details else ''
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if (i - 2) % 2 == 1:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = 15
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_len + 2

    filename = f'ACCIO_audit_log_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.xlsx'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
